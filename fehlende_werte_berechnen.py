import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl


def output_to_widget(text_widget, message):
    text_widget.insert(tk.END, message + "\n")
    text_widget.see(tk.END)
    text_widget.update_idletasks()


def calculate_and_update(file_path, text_widget):
    try:
        # Excel-Datei laden
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active  # Aktives Arbeitsblatt

        # Durch die Zeilen der Excel-Tabelle iterieren, ab der 2. Zeile
        for row in ws.iter_rows(min_row=2, min_col=8, max_col=15):  # H bis O (Spalten 8 bis 15)
            h_to_o_cells = [cell.value for cell in row]  # Werte der Zellen H bis O

            if "nachbearbeiten" in h_to_o_cells:
                # Wenn "nachbearbeiten" gefunden wird, prüfe die Zellen D bis G (Spalten 4 bis 7)
                d_to_g_values = [ws.cell(row=row[0].row, column=col).value for col in range(4, 8)]

                # Überprüfen, ob alle Zellen in D bis G Zahlenwerte enthalten
                if all(isinstance(val, (int, float)) for val in d_to_g_values):
                    max_extension_links = d_to_g_values[0]  # Wert aus Spalte D
                    max_extension_rechts = d_to_g_values[1]  # Wert aus Spalte E
                    max_flexion_links = d_to_g_values[2]  # Wert aus Spalte F
                    max_flexion_rechts = d_to_g_values[3]  # Wert aus Spalte G

                    # Berechnungen durchführen
                    seitenunterschied_extension_absolut = round(abs(max_extension_links - max_extension_rechts), 2)
                    min_extension = min(max_extension_links, max_extension_rechts)
                    max_extension = max(max_extension_links, max_extension_rechts)
                    seitenunterschied_extension_relativ = round((1 - (min_extension / max_extension)) * 100, 2)
                    seitenunterschied_flexion_absolut = round(abs(max_flexion_links - max_flexion_rechts), 2)
                    min_flexion = min(max_flexion_links, max_flexion_rechts)
                    max_flexion = max(max_flexion_links, max_flexion_rechts)
                    seitenunterschied_flexion_relativ = round((1 - (min_flexion / max_flexion)) * 100, 2)
                    verhaeltnis_flexion_extension_links = round(max_flexion_links / max_extension_links, 2)
                    verhaeltnis_flexion_extension_rechts = round(max_flexion_rechts / max_extension_rechts, 2)
                    unterschied_extension_flexion_links = round(abs(max_extension_links - max_flexion_links), 2)
                    unterschied_extension_flexion_rechts = round(abs(max_extension_rechts - max_flexion_rechts), 2)

                    # Werte in die Excel-Datei eintragen
                    current_row = row[0].row
                    ws.cell(row=current_row, column=8, value=seitenunterschied_extension_absolut)
                    ws.cell(row=current_row, column=9, value=seitenunterschied_extension_relativ)
                    ws.cell(row=current_row, column=10, value=seitenunterschied_flexion_absolut)
                    ws.cell(row=current_row, column=11, value=seitenunterschied_flexion_relativ)
                    ws.cell(row=current_row, column=12, value=verhaeltnis_flexion_extension_links)
                    ws.cell(row=current_row, column=13, value=verhaeltnis_flexion_extension_rechts)
                    ws.cell(row=current_row, column=14, value=unterschied_extension_flexion_links)
                    ws.cell(row=current_row, column=15, value=unterschied_extension_flexion_rechts)

                    output_to_widget(text_widget, f"Zeile {current_row}: Werte wurden eingetragen.")

        # Änderungen speichern
        wb.save(file_path)

        # Erfolgsnachricht
        messagebox.showinfo("Berechnung abgeschlossen",
                            f"Seitenunterschiede, Verhältnisse und Unterschiede pro Bein wurden berechnet und in {file_path} gespeichert.")

    except Exception as e:
        messagebox.showerror("Fehler", f"Fehler beim Verarbeiten der Datei: {e}")
    finally:
        wb.close()


def select_file(entry):
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        entry.delete(0, tk.END)
        entry.insert(0, file_path)


def main():
    # GUI erstellen
    root = tk.Tk()
    root.title("Isokinetik: Fehlende Werte berechnen.")

    # Datei-Pfad-Eingabe
    frame = tk.Frame(root)
    frame.pack(padx=10, pady=5)

    tk.Label(frame, text="Pfad zur Ergebnistabelle:").pack(side=tk.LEFT)
    path_entry = tk.Entry(frame, width=50)
    path_entry.pack(side=tk.LEFT, padx=(5, 0))

    # Datei durchsuchen Button
    browse_button = tk.Button(frame, text="Durchsuchen", command=lambda: select_file(path_entry))
    browse_button.pack(side=tk.LEFT, padx=(5, 0))

    # Start-Button
    start_button = tk.Button(root, text="Starten", command=lambda: calculate_and_update(path_entry.get(), output_text))
    start_button.pack(pady=5)

    # Text-Widget für die Ausgabe
    output_text = tk.Text(root, wrap="word", height=25, width=70)
    output_text.pack(padx=10, pady=10)

    # GUI ausführen
    root.mainloop()


if __name__ == "__main__":
    main()

