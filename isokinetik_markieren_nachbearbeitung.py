import openpyxl
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog, messagebox

def process_excel(file_path):
    # Definiere eine rote Füllfarbe für "nachbearbeiten"
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Definiere eine gelbe Füllfarbe für Zahlenwert-Bedingungen
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Lade die Arbeitsmappe und bearbeite sie
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Durchlaufe alle Zellen im Arbeitsblatt
        for row in sheet.iter_rows():
            for cell in row:
                # Markiere alle Zellen mit "nachbearbeiten" rot
                if cell.value == "nachbearbeiten":
                    cell.fill = red_fill
                # Prüfe, ob der Zellwert eine Zahl ist
                elif isinstance(cell.value, (int, float)):
                    # Runde die Zahl auf zwei Nachkommastellen
                    cell.value = round(cell.value, 2)

        # Gehe durch die Spalten T, U, V, W (entspricht Spalten 20 bis 23)
        for row in sheet.iter_rows(min_row=2, min_col=20, max_col=23):
            for cell in row:
                if isinstance(cell.value, str) and "-" in cell.value:  # Prüfen, ob der Zellinhalt das "-" enthält
                    try:
                        # Teile die Werte anhand des Minus und strippe Leerzeichen
                        first_val, second_val = map(str.strip, cell.value.split('-'))
                        # Ersetze Kommas durch Punkte für die Umwandlung in float
                        first_val = first_val.replace(',', '.')
                        second_val = second_val.replace(',', '.')

                        first_num = float(first_val)
                        second_num = float(second_val)

                        # Überprüfen der Bereiche
                        first_in_range = -5 <= first_num <= 5
                        second_in_range = 95 <= second_num <= 105

                        # Wenn einer oder beide Werte außerhalb ihres Bereichs liegen, markiere die Zelle gelb
                        if not first_in_range or not second_in_range:
                            cell.fill = yellow_fill  # Markiere die Zelle gelb

                        # Runde beide Werte auf zwei Nachkommastellen
                        first_num_rounded = round(first_num, 2)
                        second_num_rounded = round(second_num, 2)

                        # Setze die gerundeten Werte zurück in die Zelle (mit Kommas)
                        cell.value = f"{str(first_num_rounded).replace('.', ',')} - {str(second_num_rounded).replace('.', ',')}"

                    except ValueError:
                        # Falls der Wert kein Zahlenwert ist oder eine falsche Formatierung hat, einfach überspringen
                        pass

        # Speichere die Änderungen in der Datei
        wb.save(file_path)
        messagebox.showinfo("Erfolg", "Alle relevanten Zellen wurden entsprechend den Vorgaben markiert.")

    # Stelle sicher, dass die Datei geschlossen wird, auch bei Fehlern
    finally:
        wb.close()


def browse_file(entry):
    # Öffnet ein Dialogfenster, um eine Datei auszuwählen
    file_path = filedialog.askopenfilename(filetypes=[("Excel-Dateien", "*.xlsx")])
    if file_path:
        entry.delete(0, tk.END)  # Löscht den aktuellen Text
        entry.insert(0, file_path)  # Setzt den neuen Pfad in das Eingabefeld

def on_start_button_click(entry):
    file_path = entry.get()
    if file_path:
        try:
            process_excel(file_path)  # Verarbeitet die Excel-Datei
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler bei der Verarbeitung der Datei: {e}")
    else:
        messagebox.showwarning("Warnung", "Bitte wählen Sie eine Datei aus.")


def main():
    # Erstelle das Hauptfenster
    root = tk.Tk()
    root.title("Isokinetik: Markierungen für die Nachbearbeitung")

    # Setze die Fenstergröße (Breite x Höhe)
    root.geometry("600x300")

    # Frame für das Eingabefeld und den "Durchsuchen"-Button
    frame = tk.Frame(root)
    frame.pack(padx=10, pady=10)

    entry_label = tk.Label(frame, text="Wähle die Ergebnistabelle aus:")
    entry_label.grid(row=0, column=0, sticky="w")

    entry = tk.Entry(frame, width=50)
    entry.grid(row=0, column=1)

    browse_button = tk.Button(frame, text="Durchsuchen", command=lambda: browse_file(entry))
    browse_button.grid(row=0, column=2, padx=5)

    # Start-Button
    start_button = tk.Button(root, text="Start", command=lambda: on_start_button_click(entry))
    start_button.pack(pady=10)

    # Starte die GUI
    root.mainloop()

# Überprüfe, ob das Skript direkt ausgeführt wird, und starte dann die GUI
if __name__ == "__main__":
    main()

